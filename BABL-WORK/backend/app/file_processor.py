import pandas as pd
import logging
import re
from typing import Dict, List, Tuple, Optional
from datetime import datetime
import uuid

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FileProcessor:
    def __init__(self):
        self.id_counter = {}
        
    def generate_id(self, pharmacy_name: str, row_index: int) -> str:
        """
        Generate ID using the same logic as tasks_enhanced.py
        Uses full pharmacy name for both facility and location parts
        """
        from app.tasks_enhanced import generate_id
        return generate_id(pharmacy_name, pharmacy_name, row_index, {})
    
    def validate_invoice_columns(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Validate that invoice file has required columns
        Expected: Pharmacy_Name | Product | Quantity | Amount
        """
        required_columns = ['Pharmacy_Name', 'Product', 'Quantity', 'Amount']
        missing_columns = []
        
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        return len(missing_columns) == 0, missing_columns
    
    def validate_master_columns(self, df: pd.DataFrame) -> Tuple[bool, List[str]]:
        """
        Validate that master file has required columns
        Expected: REP_Names | Doctor_Names | Doctor_ID | Pharmacy_Names | Pharmacy_ID | 
                 Product_Names | Product_ID | Product_Price | HQ | AREA
        """
        required_columns = [
            'REP_Names', 'Doctor_Names', 'Doctor_ID', 'Pharmacy_Names', 'Pharmacy_ID',
            'Product_Names', 'Product_ID', 'Product_Price', 'HQ', 'AREA'
        ]
        missing_columns = []
        
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        return len(missing_columns) == 0, missing_columns
    
    def process_invoice_file(self, file_path: str, user_id: int = 1) -> Dict:
        """
        Process invoice file and generate pharmacy IDs, storing directly in database
        Ensures ALL rows are read and processed correctly
        
        Args:
            file_path: Path to the invoice Excel file
            user_id: ID of the user uploading the file (default: 1)
        """
        try:
            from app.database import get_db, Invoice
            
            # Read Excel file - ensure we read ALL rows including empty ones initially
            # Then filter out completely empty rows
            logger.info(f"Reading Excel file: {file_path}")
            
            # Read with explicit parameters to ensure we get all data
            # Don't use dtype=str as it causes issues with numeric columns
            # Instead, read normally and handle NaN values properly
            # Use header=0 to ensure first row is treated as header
            df = pd.read_excel(
                file_path, 
                engine='openpyxl', 
                keep_default_na=False,
                na_values=['', ' ', 'NULL', 'null', 'None', 'none'],
                header=0  # Explicitly set first row as header
            )
            
            # Verify we read the file correctly by checking total rows
            # Open the file with openpyxl directly to count actual rows
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, read_only=True, data_only=True)
                ws = wb.active
                # Count non-empty rows (skip header)
                actual_row_count = sum(1 for row in ws.iter_rows(min_row=2) if any(cell.value for cell in row))
                wb.close()
                logger.info(f"Excel file has {actual_row_count} data rows (excluding header) according to openpyxl")
                logger.info(f"Pandas read {len(df)} rows")
                if abs(actual_row_count - len(df)) > 5:  # Allow small difference for completely empty rows
                    logger.warning(f"Row count mismatch! Excel has {actual_row_count} rows, pandas read {len(df)} rows")
            except Exception as e:
                logger.warning(f"Could not verify row count with openpyxl: {str(e)}")
            
            # Log initial file reading stats
            logger.info(f"Initial file read: {len(df)} total rows, {len(df.columns)} columns")
            logger.info(f"Columns found: {list(df.columns)}")
            
            # Count rows by pharmacy name BEFORE filtering (for debugging)
            if 'Pharmacy_Name' in df.columns:
                try:
                    pharmacy_counts = df['Pharmacy_Name'].astype(str).str.strip().value_counts()
                    # Filter Ace Care using boolean indexing properly
                    ace_care_mask = pharmacy_counts.index.str.contains('ace care', case=False, na=False)
                    ace_care_in_file = pharmacy_counts[ace_care_mask]
                    if len(ace_care_in_file) > 0:
                        logger.info(f"Ace Care rows in file (before filtering): {ace_care_in_file.to_dict()}")
                    logger.info(f"Top 10 pharmacies by row count: {pharmacy_counts.head(10).to_dict()}")
                except Exception as e:
                    logger.warning(f"Error counting pharmacies: {str(e)}")
            
            # Remove completely empty rows (where all key columns are empty)
            # But keep rows where at least one key field has data
            required_cols = ['Pharmacy_Name', 'Product', 'Quantity', 'Amount']
            before_filter = len(df)
            
            # Filter out rows where ALL required columns are empty/NaN
            df = df.dropna(subset=required_cols, how='all')
            
            # Also remove rows where Pharmacy_Name is empty (most critical field)
            # Create boolean mask to avoid Series boolean ambiguity
            pharmacy_name_series = df['Pharmacy_Name'].astype(str).str.strip()
            non_empty_mask = (pharmacy_name_series != '') & (pharmacy_name_series.str.lower() != 'nan')
            df = df[non_empty_mask]
            
            after_filter = len(df)
            logger.info(f"After filtering empty rows: {before_filter} -> {after_filter} rows")
            
            # Count rows by pharmacy name AFTER filtering
            if 'Pharmacy_Name' in df.columns:
                try:
                    pharmacy_counts_after = df['Pharmacy_Name'].astype(str).str.strip().value_counts()
                    # Filter Ace Care using boolean indexing properly
                    ace_care_mask_after = pharmacy_counts_after.index.str.contains('ace care', case=False, na=False)
                    ace_care_after = pharmacy_counts_after[ace_care_mask_after]
                    if len(ace_care_after) > 0:
                        logger.info(f"Ace Care rows after filtering: {ace_care_after.to_dict()}")
                        total_ace_care_rows = ace_care_after.sum()
                        logger.info(f"Total Ace Care rows after filtering: {total_ace_care_rows}")
                except Exception as e:
                    logger.warning(f"Error counting pharmacies after filtering: {str(e)}")
            
            # Validate columns
            is_valid, missing_cols = self.validate_invoice_columns(df)
            if not is_valid:
                return {
                    "success": False,
                    "error": f"Missing required columns: {', '.join(missing_cols)}",
                    "processed_rows": 0
                }
            
            # Get database session
            db = next(get_db())
            
            try:
                # Process each row and generate IDs
                processed_data = []
                unmatched_pharmacies = []
                ace_care_count = 0
                ace_care_total = 0.0
                
                # Reset index to ensure we have sequential row numbers
                df = df.reset_index(drop=True)
                
                logger.info(f"Processing {len(df)} rows from file")
                
                for index, row in df.iterrows():
                    # Get values with proper handling of NaN/empty
                    pharmacy_name = str(row['Pharmacy_Name']).strip() if pd.notna(row['Pharmacy_Name']) else ''
                    product = str(row['Product']).strip() if pd.notna(row['Product']) else ''
                    
                    # Handle quantity and amount - convert to numeric, default to 0 if invalid
                    try:
                        quantity = float(row['Quantity']) if pd.notna(row['Quantity']) and str(row['Quantity']).strip() != '' else 0.0
                    except (ValueError, TypeError):
                        quantity = 0.0
                    
                    try:
                        amount = float(row['Amount']) if pd.notna(row['Amount']) and str(row['Amount']).strip() != '' else 0.0
                    except (ValueError, TypeError):
                        amount = 0.0
                    
                    # Skip rows with empty pharmacy name (shouldn't happen after filtering, but double-check)
                    if not pharmacy_name or pharmacy_name.lower() == 'nan':
                        logger.warning(f"Row {index + 2}: Skipping row with empty pharmacy name")
                        continue
                    
                    # Track Ace Care for debugging
                    if 'ace care' in pharmacy_name.lower():
                        ace_care_count += 1
                        ace_care_total += amount
                        if ace_care_count <= 5 or ace_care_count % 10 == 0:
                            logger.info(f"Ace Care row {ace_care_count}: {pharmacy_name}, Product: {product}, Qty: {quantity}, Amount: {amount}")
                    
                    # Use full pharmacy name for both facility and location (no splitting)
                    # Generate ID using the same logic as tasks_enhanced.py
                    from app.tasks_enhanced import generate_id
                    generated_id = generate_id(pharmacy_name, pharmacy_name, index, {})
                    
                    # Store in database
                    if generated_id != 'INVALID':
                        invoice_record = Invoice(
                            pharmacy_id=generated_id.replace('-', '_'),
                            pharmacy_name=pharmacy_name,
                            product=product,
                            quantity=int(quantity) if quantity > 0 else 0,
                            amount=float(amount),
                            user_id=user_id,  # Use provided user_id
                            invoice_date=datetime.now()  # Set invoice date
                        )
                        db.add(invoice_record)
                    
                    processed_row = {
                        "original_pharmacy_name": pharmacy_name,
                        "generated_id": generated_id,
                        "product": product,
                        "quantity": quantity,
                        "amount": amount,
                        "row_index": index + 2  # +2 because Excel is 1-indexed and has header
                    }
                    
                    processed_data.append(processed_row)
                    
                    # Track unmatched pharmacies (those with INVALID IDs)
                    if generated_id == "INVALID":
                        unmatched_pharmacies.append({
                            "pharmacy_name": pharmacy_name,
                            "generated_id": generated_id,
                            "row_index": index + 2,
                            "reason": "Invalid pharmacy name"
                        })
                
                # Log Ace Care summary
                logger.info(f"Ace Care summary: {ace_care_count} rows found, Total amount: {ace_care_total:.2f}")
                
                # Commit to database
                db.commit()
                logger.info(f"Committed {len(processed_data)} invoice records to database")
                
                # Verify what was actually stored for Ace Care
                ace_care_in_db = db.query(Invoice).filter(Invoice.pharmacy_name.ilike('%ace care%')).all()
                ace_care_db_count = len(ace_care_in_db)
                ace_care_db_total = sum(float(inv.amount) for inv in ace_care_in_db)
                logger.info(f"Ace Care in database: {ace_care_db_count} records, Total: {ace_care_db_total:.2f}")
                
                # Create pharmacy summary for debugging
                pharmacy_summary = {}
                for row in processed_data:
                    pharm_name = row['original_pharmacy_name']
                    if pharm_name not in pharmacy_summary:
                        pharmacy_summary[pharm_name] = {'count': 0, 'total_amount': 0.0}
                    pharmacy_summary[pharm_name]['count'] += 1
                    pharmacy_summary[pharm_name]['total_amount'] += row['amount']
                
                # Log top pharmacies
                sorted_pharmacies = sorted(pharmacy_summary.items(), key=lambda x: x[1]['total_amount'], reverse=True)
                logger.info(f"Top 10 pharmacies by revenue: {[(name, data['count'], data['total_amount']) for name, data in sorted_pharmacies[:10]]}")
                
                # Do not create a RecentUpload here; analysis endpoint will create one
                
            except Exception as e:
                db.rollback()
                logger.error(f"Error processing rows: {str(e)}", exc_info=True)
                raise e
            finally:
                db.close()
            
            # Create pharmacy breakdown for response
            pharmacy_breakdown = {}
            for row in processed_data:
                pharm_name = row['original_pharmacy_name']
                if pharm_name not in pharmacy_breakdown:
                    pharmacy_breakdown[pharm_name] = {'rows': 0, 'total_amount': 0.0}
                pharmacy_breakdown[pharm_name]['rows'] += 1
                pharmacy_breakdown[pharm_name]['total_amount'] += row['amount']
            
            return {
                "success": True,
                "processed_rows": len(processed_data),
                "data": processed_data,
                "unmatched_pharmacies": unmatched_pharmacies,
                "summary": {
                    "total_rows": len(df),
                    "valid_rows": len([d for d in processed_data if d["generated_id"] != "INVALID"]),
                    "invalid_rows": len(unmatched_pharmacies),
                    "unique_pharmacies": len(set(d["generated_id"] for d in processed_data if d["generated_id"] != "INVALID")),
                    "ace_care_rows": ace_care_count,
                    "ace_care_total": ace_care_total
                },
                "pharmacy_breakdown": {k: {"rows": v["rows"], "total_amount": round(v["total_amount"], 2)} for k, v in sorted(pharmacy_breakdown.items(), key=lambda x: x[1]["total_amount"], reverse=True)[:20]}
            }
            
        except Exception as e:
            logger.error(f"Error processing invoice file: {str(e)}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "processed_rows": 0
            }
    
    def process_master_file(self, file_path: str) -> Dict:
        """
        Process master file and validate data, storing directly in database
        """
        try:
            from app.database import get_db, MasterMapping
            
            # Read Excel file
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                logger.info(f"Read Excel file: {len(df)} rows, {len(df.columns)} columns")
            except Exception as read_error:
                logger.error(f"Error reading Excel file: {str(read_error)}", exc_info=True)
                return {
                    "success": False,
                    "error": f"Error reading Excel file: {str(read_error)}",
                    "processed_rows": 0
                }
            
            # Validate columns
            is_valid, missing_cols = self.validate_master_columns(df)
            if not is_valid:
                return {
                    "success": False,
                    "error": f"Missing required columns: {', '.join(missing_cols)}",
                    "processed_rows": 0
                }
            
            # Get database session
            db = next(get_db())
            
            try:
                # Process and validate data
                processed_data = []
                validation_errors = []
                
                for index, row in df.iterrows():
                    # Truncate to safe limits (matching old DB schema as fallback)
                    # Migration will update DB, but this ensures compatibility
                    processed_row = {
                        "rep_name": (str(row['REP_Names']).strip()[:200] if pd.notna(row['REP_Names']) else ""),
                        "doctor_name": (str(row['Doctor_Names']).strip()[:200] if pd.notna(row['Doctor_Names']) else ""),
                        "doctor_id": (str(row['Doctor_ID']).strip()[:100] if pd.notna(row['Doctor_ID']) else ""),
                        "pharmacy_name": (str(row['Pharmacy_Names']).strip()[:500] if pd.notna(row['Pharmacy_Names']) else ""),
                        "pharmacy_id": (str(row['Pharmacy_ID']).strip()[:100] if pd.notna(row['Pharmacy_ID']) else ""),
                        "product_name": (str(row['Product_Names']).strip()[:300] if pd.notna(row['Product_Names']) else ""),
                        "product_id": (str(row['Product_ID']).strip()[:100] if pd.notna(row['Product_ID']) else ""),
                        "product_price": float(row['Product_Price']) if pd.notna(row['Product_Price']) else 0.0,
                        "hq": (str(row['HQ']).strip()[:100] if pd.notna(row['HQ']) else ""),
                        "area": (str(row['AREA']).strip()[:200] if pd.notna(row['AREA']) else ""),
                        "row_index": index + 2
                    }
                    
                    # Additional safety: ensure all strings are within old limits as absolute fallback
                    # This prevents errors if migration hasn't run yet
                    processed_row["rep_name"] = processed_row["rep_name"][:200]
                    processed_row["doctor_name"] = processed_row["doctor_name"][:200]
                    processed_row["doctor_id"] = processed_row["doctor_id"][:100]
                    processed_row["pharmacy_name"] = processed_row["pharmacy_name"][:500]
                    processed_row["pharmacy_id"] = processed_row["pharmacy_id"][:100]
                    processed_row["product_name"] = processed_row["product_name"][:300]
                    processed_row["product_id"] = processed_row["product_id"][:100] if processed_row["product_id"] else None
                    processed_row["hq"] = processed_row["hq"][:100] if processed_row["hq"] else ""
                    processed_row["area"] = processed_row["area"][:200] if processed_row["area"] else ""
                    
                    # Validate required fields
                    if not processed_row["pharmacy_name"]:
                        validation_errors.append({
                            "row": index + 2,
                            "field": "Pharmacy_Names",
                            "error": "Pharmacy name is required"
                        })
                    
                    if not processed_row["pharmacy_id"]:
                        validation_errors.append({
                            "row": index + 2,
                            "field": "Pharmacy_ID",
                            "error": "Pharmacy ID is required"
                        })
                    
                    if processed_row["product_price"] <= 0:
                        validation_errors.append({
                            "row": index + 2,
                            "field": "Product_Price",
                            "error": "Product price must be greater than 0"
                        })
                    
                    # Store in database if valid
                    if not validation_errors or all(error["row"] != index + 2 for error in validation_errors):
                        # CRITICAL: Truncate to OLD database limits (50) as safety fallback
                        # This ensures compatibility even if migration hasn't run
                        # After migration succeeds, these will be updated to larger sizes
                        safe_pharmacy_id = str(processed_row["pharmacy_id"]).replace('-', '_')[:50]
                        safe_pharmacy_names = str(processed_row["pharmacy_name"])[:50]
                        safe_product_names = str(processed_row["product_name"])[:50]
                        safe_product_id = str(processed_row["product_id"])[:50] if processed_row["product_id"] else None
                        safe_doctor_names = str(processed_row["doctor_name"])[:50]
                        safe_doctor_id = str(processed_row["doctor_id"])[:50]
                        safe_rep_names = str(processed_row["rep_name"])[:50]
                        safe_hq = str(processed_row["hq"])[:50] if processed_row["hq"] else ""
                        safe_area = str(processed_row["area"])[:50] if processed_row["area"] else ""
                        
                        master_record = MasterMapping(
                            pharmacy_id=safe_pharmacy_id,
                            pharmacy_names=safe_pharmacy_names,
                            product_names=safe_product_names,
                            product_id=safe_product_id,
                            product_price=processed_row["product_price"],
                            doctor_names=safe_doctor_names,
                            doctor_id=safe_doctor_id,
                            rep_names=safe_rep_names,
                            hq=safe_hq,
                            area=safe_area
                        )
                        db.add(master_record)
                    
                    processed_data.append(processed_row)
                
                # Commit to database
                try:
                    db.commit()
                    logger.info(f"Committed {len(processed_data)} master records to database")
                except Exception as commit_error:
                    db.rollback()
                    logger.error(f"Error committing master data: {str(commit_error)}", exc_info=True)
                    raise Exception(f"Database commit failed: {str(commit_error)}")
                
                # Do not create a RecentUpload here; analysis endpoint will create one
                
            except Exception as e:
                db.rollback()
                logger.error(f"Error processing master data rows: {str(e)}", exc_info=True)
                raise e
            finally:
                db.close()
            
            return {
                "success": True,
                "processed_rows": len(processed_data),
                "data": processed_data,
                "validation_errors": validation_errors,
                "summary": {
                    "total_rows": len(df),
                    "valid_rows": len(processed_data) - len(validation_errors),
                    "error_rows": len(validation_errors),
                    "unique_pharmacies": len(set(d["pharmacy_id"] for d in processed_data if d["pharmacy_id"])),
                    "unique_products": len(set(d["product_id"] for d in processed_data if d["product_id"]))
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing master file: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "processed_rows": 0
            }
    
    def match_invoice_with_master(self, invoice_data: List[Dict], master_data: List[Dict]) -> Dict:
        """
        Match invoice data with master data using STRICT Pharmacy + Product matching
        """
        # Create master lookup by pharmacy_id + product
        master_lookup = {}
        
        # Use the standard normalize_product_name from tasks_enhanced for consistency
        from app.tasks_enhanced import normalize_product_name
        
        for master_row in master_data:
            # Normalize product name for matching using standard function
            normalized_product = normalize_product_name(master_row["product_name"])
            key = f"{master_row['pharmacy_id']}|{normalized_product}"
            master_lookup[key] = master_row
        
        matched_data = []
        unmatched_invoices = []
        
        for invoice_row in invoice_data:
            generated_id = invoice_row["generated_id"]
            
            # Normalize ID for matching (replace - with _)
            normalized_id = generated_id.replace("-", "_")
            
            # Normalize product name for matching using standard function
            normalized_product = normalize_product_name(invoice_row["product"])
            
            # Create composite key for lookup
            lookup_key = f"{normalized_id}|{normalized_product}"
            
            # Try to find exact match for BOTH pharmacy and product
            matched_master = master_lookup.get(lookup_key)
            
            if matched_master:
                # Calculate revenue: Quantity × Master.Product_Price
                quantity = invoice_row["quantity"]
                product_price = matched_master["product_price"]
                calculated_amount = quantity * product_price
                
                matched_row = {
                    **invoice_row,
                    "master_pharmacy_id": matched_master["pharmacy_id"],
                    "master_pharmacy_name": matched_master["pharmacy_name"],
                    "doctor_name": matched_master["doctor_name"],
                    "doctor_id": matched_master["doctor_id"],
                    "rep_name": matched_master["rep_name"],
                    "product_name": matched_master["product_name"],
                    "product_id": matched_master["product_id"],
                    "product_price": product_price,
                    "hq": matched_master["hq"],
                    "area": matched_master["area"],
                    "calculated_amount": calculated_amount,
                    "match_status": "matched"
                }
                matched_data.append(matched_row)
            else:
                # No match for this pharmacy+product combination
                unmatched_invoices.append({
                    **invoice_row,
                    "match_status": "unmatched",
                    "reason": "No matching pharmacy+product combination found in master data"
                })
        
        return {
            "matched_data": matched_data,
            "unmatched_invoices": unmatched_invoices,
            "summary": {
                "total_invoices": len(invoice_data),
                "matched_count": len(matched_data),
                "unmatched_count": len(unmatched_invoices),
                "match_rate": len(matched_data) / len(invoice_data) * 100 if invoice_data else 0
            }
        }
    
    # REMOVED: _normalize_product_name - now using standard normalize_product_name from tasks_enhanced
    # This eliminates duplicate normalization logic and ensures consistent matching
